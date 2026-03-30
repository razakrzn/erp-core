from __future__ import annotations

import io
import logging
from pathlib import Path
import math
from contextlib import contextmanager

import ezdxf  # type: ignore[import-untyped]
import matplotlib
import numpy as np  # type: ignore[import-untyped]
from openpyxl import Workbook
from ezdxf.addons import odafc  # type: ignore[import-untyped]
from ezdxf.addons.drawing import Frontend, RenderContext
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend

matplotlib.use("Agg")
from matplotlib.backends.backend_pdf import PdfPages  # noqa: E402
import matplotlib.pyplot as plt  # noqa: E402
from rectpack import newPacker  # type: ignore[import-untyped]  # noqa: E402
from reportlab.lib.pagesizes import A4  # noqa: E402
from reportlab.pdfgen import canvas  # noqa: E402
from shapely.geometry import Polygon  # type: ignore[import-untyped]  # noqa: E402

logger = logging.getLogger(__name__)
_CAD_RENDER_NOISE_SNIPPETS = (
    "copy process ignored ACDB_BLOCKREPRESENTATION_DATA",
    "relative point size is not supported",
)


class _CadRenderNoiseFilter(logging.Filter):
    def filter(self, record: logging.LogRecord) -> bool:
        message = record.getMessage()
        return not any(snippet in message for snippet in _CAD_RENDER_NOISE_SNIPPETS)


@contextmanager
def _suppress_known_cad_render_noise():
    noise_filter = _CadRenderNoiseFilter()
    root_logger = logging.getLogger()
    ezdxf_logger = logging.getLogger("ezdxf")
    root_logger.addFilter(noise_filter)
    ezdxf_logger.addFilter(noise_filter)
    try:
        yield
    finally:
        root_logger.removeFilter(noise_filter)
        ezdxf_logger.removeFilter(noise_filter)


def _round_dim(value: float) -> float:
    return round(max(float(value), 0.0), 3)


def _guess_scale_factor(parts: list[dict]) -> float:
    """
    Makes an educated guess about the unit based on the maximum dimension
    found in the extracted parts. Used when $INSUNITS is unitless/unspecified.
    """
    if not parts:
        return 1.0

    max_dim = max(max(p["width"], p["height"]) for p in parts)

    # Heuristic Rule 1: If the largest part is very small (e.g., < 15 units)
    # It's highly probable the drawing was done in Meters.
    if max_dim < 15.0:
        return 1000.0  # Scale Meters to Millimeters

    # Heuristic Rule 2: If the largest part is medium-sized (e.g., < 150 units)
    # 150 inches is ~3.8 meters, typical max for sheet materials.
    # Likely the drawing was done in Inches.
    elif max_dim < 150.0:
        return 25.4  # Scale Inches to Millimeters

    # Default: Large values (e.g., 500, 1200, 3000) suggest Millimeters.
    return 1.0


def _get_scale_factor_to_mm(doc, parts: list[dict]) -> float:
    """
    Determines the scale factor to convert DXF units to millimeters.
    Checks the $INSUNITS header first. If missing or unitless, falls back to guessing.
    """
    insunits = doc.header.get("$INSUNITS", 0)

    # Standard AutoCAD $INSUNITS codes to mm
    if insunits == 4:
        return 1.0  # Millimeters
    elif insunits == 5:
        return 10.0  # Centimeters
    elif insunits == 6:
        return 1000.0  # Meters
    elif insunits == 1:
        return 25.4  # Inches
    elif insunits == 2:
        return 304.8  # Feet

    # Unitless (0) or unrecognized code: use smart guess
    return _guess_scale_factor(parts)


def _extract_parts_from_doc(doc, min_dimension: float = 20.0) -> list[dict]:
    """
    Extract parts from DXF/DWG. Ignores parts smaller than min_dimension (mm)
    to filter out junk artifacts.
    """
    parts = []
    msp = doc.modelspace()

    for entity in msp.query("LWPOLYLINE"):
        if not entity.closed:
            continue

        points = [(p[0], p[1]) for p in entity.get_points()]
        if len(points) < 3:
            continue

        polygon = Polygon(points)
        if not polygon.is_valid or polygon.area <= 0:
            continue

        min_rect = polygon.minimum_rotated_rectangle
        coords = list(min_rect.exterior.coords)
        edge1 = math.dist(coords[0], coords[1])
        edge2 = math.dist(coords[1], coords[2])

        width = _round_dim(min(edge1, edge2))
        height = _round_dim(max(edge1, edge2))
        if width > min_dimension and height > min_dimension:
            parts.append({"width": width, "height": height, "quantity": 1})

    for entity in msp.query("CIRCLE"):
        diameter = _round_dim(entity.dxf.radius * 2)
        if diameter > min_dimension:
            parts.append({"width": diameter, "height": diameter, "quantity": 1})

    for entity in msp.query("SPLINE"):
        try:
            points = list(entity.flattening(0.5))
        except Exception:
            continue
        if len(points) < 2:
            continue
        minx, maxx = min(p[0] for p in points), max(p[0] for p in points)
        miny, maxy = min(p[1] for p in points), max(p[1] for p in points)
        bounds_polygon = Polygon(
            [
                (minx, miny),
                (maxx, miny),
                (maxx, maxy),
                (minx, maxy),
            ]
        )
        if not bounds_polygon.is_valid or bounds_polygon.area <= 0:
            continue
        min_rect = bounds_polygon.minimum_rotated_rectangle
        rect_coords = list(min_rect.exterior.coords)
        edge1 = math.dist(rect_coords[0], rect_coords[1])
        edge2 = math.dist(rect_coords[1], rect_coords[2])
        width = _round_dim(min(edge1, edge2))
        height = _round_dim(max(edge1, edge2))
        if width > min_dimension and height > min_dimension:
            parts.append({"width": width, "height": height, "quantity": 1})

    return parts


def _load_doc(file_path: Path):
    suffix = file_path.suffix.lower()
    if suffix == ".dxf":
        return ezdxf.readfile(file_path)

    if suffix == ".dwg":
        try:
            return odafc.readfile(file_path)
        except Exception as exc:  # pragma: no cover - depends on ODA converter
            raise ValueError(
                "DWG parsing requires ODA File Converter support in ezdxf. "
                "Install/configure ODA converter or upload DXF."
            ) from exc

    raise ValueError("Unsupported CAD file format. Upload .dxf or .dwg")


def _build_default_sheets() -> list[dict]:
    return [{"name": "Default Sheet", "width": 3000.0, "height": 1500.0, "quantity": 50}]


def run_cutting_optimization(
    file_path: str,
    stock_sheets: list[dict] | None = None,
    kerf: float = 0.0,
    min_dimension: float = 20.0,
) -> dict:
    path = Path(file_path)
    doc = _load_doc(path)
    parts = _extract_parts_from_doc(doc, min_dimension=min_dimension)

    if not parts:
        return {
            "parts": [],
            "used_bins": [],
            "placements": [],
            "unplaced_parts": [],
            "oversized_parts": [],
            "summary": {
                "total_parts": 0,
                "placed_parts": 0,
                "unplaced_parts": 0,
                "oversized_parts": 0,
                "utilization_percent": 0.0,
            },
        }

    scale_factor = _get_scale_factor_to_mm(doc, parts)
    if scale_factor != 1.0:
        for part in parts:
            part["width"] = _round_dim(part["width"] * scale_factor)
            part["height"] = _round_dim(part["height"] * scale_factor)

    normalized_sheets = stock_sheets or _build_default_sheets()

    # Filter out parts that exceed max sheet size (cannot be packed)
    # Use best-available sheet dimensions (max width/height across all sheet types)
    max_sheet_width = max(_round_dim(s["width"]) for s in normalized_sheets)
    max_sheet_height = max(_round_dim(s["height"]) for s in normalized_sheets)
    sheet_max = max(max_sheet_width, max_sheet_height)
    sheet_min = min(max_sheet_width, max_sheet_height)

    valid_parts = []
    oversized_parts = []
    for part in parts:
        part_max = max(part["width"], part["height"])
        part_min = min(part["width"], part["height"])
        if part_max > sheet_max or part_min > sheet_min:
            oversized_parts.append(part)
        else:
            valid_parts.append(part)

    packer = newPacker(rotation=True)

    expanded_parts = []
    for idx, part in enumerate(valid_parts):
        quantity = int(part.get("quantity", 1) or 1)
        width_with_kerf = _round_dim(part["width"] + kerf)
        height_with_kerf = _round_dim(part["height"] + kerf)
        for part_index in range(quantity):
            rid = f"part-{idx + 1}-{part_index + 1}"
            expanded_parts.append(
                {
                    "rid": rid,
                    "width": width_with_kerf,
                    "height": height_with_kerf,
                    "original_w": _round_dim(part["width"]),
                    "original_h": _round_dim(part["height"]),
                }
            )
            packer.add_rect(width_with_kerf, height_with_kerf, rid=rid)

    # Large pool of sheets per type so packer can determine exactly how many are needed
    sheet_dimensions = []
    for sheet_idx, sheet in enumerate(normalized_sheets):
        quantity = 100
        width = _round_dim(sheet["width"])
        height = _round_dim(sheet["height"])
        for q in range(quantity):
            bin_id = f"sheet-{sheet_idx + 1}-{q + 1}"
            sheet_dimensions.append({"sheet": bin_id, "width": width, "height": height})
            packer.add_bin(width, height, bid=bin_id)

    packer.pack()

    placed = []
    used_bin_ids = set()
    for sheet_index, x, y, packed_width, packed_height, rid in packer.rect_list():
        bin_id = sheet_dimensions[sheet_index]["sheet"]
        used_bin_ids.add(bin_id)
        placed.append(
            {
                "part_id": rid,
                "x": _round_dim(x),
                "y": _round_dim(y),
                "width": _round_dim(packed_width - kerf),
                "height": _round_dim(packed_height - kerf),
                "sheet": bin_id,
            }
        )

    placed_ids = {item["part_id"] for item in placed}
    unplaced = [item for item in expanded_parts if item["rid"] not in placed_ids]

    used_bins = [item for item in sheet_dimensions if item["sheet"] in used_bin_ids]
    placed_area = float(np.sum([p["width"] * p["height"] for p in placed])) if placed else 0.0
    used_bin_area = float(np.sum([b["width"] * b["height"] for b in used_bins])) if used_bins else 0.0
    utilization = round((placed_area / used_bin_area) * 100, 2) if used_bin_area else 0.0

    return {
        "parts": valid_parts,
        "used_bins": used_bins,
        "placements": placed,
        "unplaced_parts": unplaced,
        "oversized_parts": oversized_parts,
        "summary": {
            "total_parts": len(expanded_parts),
            "placed_parts": len(placed),
            "unplaced_parts": len(unplaced),
            "oversized_parts": len(oversized_parts),
            "utilization_percent": utilization,
        },
    }


def _sanitize_stem(value: str) -> str:
    stem = (value or "cutting-job").strip().replace(" ", "-")
    safe = "".join(char for char in stem if char.isalnum() or char in {"-", "_"})
    return safe or "cutting-job"


def _layout_sequence(doc) -> list:
    layouts = [doc.modelspace()]
    layouts.extend(layout for layout in doc.layouts if layout.name.lower() != "model")
    return layouts


def _render_doc_to_pdf_bytes(doc, source_name: str) -> bytes:
    buffer = io.BytesIO()
    render_context = RenderContext(doc)
    with _suppress_known_cad_render_noise():
        with PdfPages(buffer) as pdf:
            for layout in _layout_sequence(doc):
                fig = plt.figure(figsize=(11.69, 8.27))
                ax = fig.add_axes((0.02, 0.02, 0.96, 0.93))
                ax.set_axis_off()
                backend = MatplotlibBackend(ax)
                Frontend(render_context, backend).draw_layout(layout, finalize=True)
                fig.suptitle(f"{source_name} - {layout.name}", fontsize=10)
                pdf.savefig(fig, bbox_inches="tight", dpi=150)
                plt.close(fig)
    return buffer.getvalue()


def _remove_text_entities(doc) -> None:
    for layout in _layout_sequence(doc):
        for entity in list(layout.query("TEXT MTEXT ATTRIB ATTDEF")):
            layout.delete_entity(entity)


def build_cad_pdf(file_path: str, source_name: str) -> bytes:
    """
    Render complete CAD drawing (model + paper layouts) into a multi-page PDF.
    """
    doc = _load_doc(Path(file_path))
    try:
        # Vector PDF output keeps quality high while reducing file size pressure.
        return _render_doc_to_pdf_bytes(doc, source_name)
    except Exception as exc:
        error_text = str(exc).lower()
        if "font" not in error_text:
            raise
        logger.warning(
            "CAD PDF rendering hit a font error for %s; retrying without text entities.",
            source_name,
            exc_info=True,
        )
        fallback_doc = _load_doc(Path(file_path))
        _remove_text_entities(fallback_doc)
        return _render_doc_to_pdf_bytes(fallback_doc, source_name)


def _write_wrapped_lines(pdf: canvas.Canvas, lines: list[str], top_y: float) -> None:
    y = top_y
    for line in lines:
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y = A4[1] - 50
        pdf.drawString(40, y, line)
        y -= 14


def build_cutlist_pdf(result: dict, source_name: str) -> bytes:
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    pdf.setTitle(f"Cutlist-{source_name}")
    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, A4[1] - 40, f"Cutlist - {source_name}")
    pdf.setFont("Helvetica", 10)

    summary = result.get("summary", {})
    lines = [
        f"Total parts: {summary.get('total_parts', 0)}",
        f"Placed parts: {summary.get('placed_parts', 0)}",
        f"Unplaced parts: {summary.get('unplaced_parts', 0)}",
        f"Oversized parts: {summary.get('oversized_parts', 0)}",
        f"Utilization: {summary.get('utilization_percent', 0.0)}%",
        "",
        "Placements:",
    ]
    for item in result.get("placements", []):
        lines.append(
            " - "
            f"{item.get('part_id')} sheet={item.get('sheet')} "
            f"x={item.get('x')} y={item.get('y')} "
            f"w={item.get('width')} h={item.get('height')}"
        )
    _write_wrapped_lines(pdf, lines, A4[1] - 65)
    pdf.save()
    return buffer.getvalue()


def build_cutlist_xlsx(result: dict) -> bytes:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    summary_ws.append(["metric", "value"])
    summary = result.get("summary", {})
    summary_ws.append(["total_parts", summary.get("total_parts", 0)])
    summary_ws.append(["placed_parts", summary.get("placed_parts", 0)])
    summary_ws.append(["unplaced_parts", summary.get("unplaced_parts", 0)])
    summary_ws.append(["oversized_parts", summary.get("oversized_parts", 0)])
    summary_ws.append(["utilization_percent", summary.get("utilization_percent", 0.0)])

    placement_ws = workbook.create_sheet("Placements")
    placement_ws.append(["part_id", "sheet", "x", "y", "width", "height"])
    for row in result.get("placements", []):
        placement_ws.append(
            [
                row.get("part_id"),
                row.get("sheet"),
                row.get("x"),
                row.get("y"),
                row.get("width"),
                row.get("height"),
            ]
        )

    unplaced_ws = workbook.create_sheet("Unplaced")
    unplaced_ws.append(["part_id", "width", "height"])
    for row in result.get("unplaced_parts", []):
        unplaced_ws.append([row.get("rid"), row.get("original_w"), row.get("original_h")])

    oversized_ws = workbook.create_sheet("Oversized")
    oversized_ws.append(["width", "height", "quantity"])
    for row in result.get("oversized_parts", []):
        oversized_ws.append([row.get("width"), row.get("height"), row.get("quantity", 1)])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_cutting_job_artifacts(result: dict, source_name: str, source_file_path: str) -> dict:
    stem = _sanitize_stem(Path(source_name).stem)
    return {
        "cad_pdf_name": f"{stem}.pdf",
        "cad_pdf_bytes": build_cad_pdf(source_file_path, source_name),
        "cutlist_pdf_name": f"{stem}-cutlist.pdf",
        "cutlist_pdf_bytes": build_cutlist_pdf(result, source_name),
        "cutlist_xlsx_name": f"{stem}-cutlist.xlsx",
        "cutlist_xlsx_bytes": build_cutlist_xlsx(result),
    }
